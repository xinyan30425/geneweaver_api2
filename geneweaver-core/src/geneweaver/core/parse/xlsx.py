"""
Parse Excel files for use by the client library.
这个模块提供了解析Excel文件的功能，以供客户端库使用。

功能说明：

- `get_sheet_names`: 获取Excel文件中所有工作表的名称。
  - 参数 `file_path`：Excel文件的路径。
  - 返回值：代表文件中所有工作表名称的字符串列表。

- `get_sheet`: 加载工作簿并返回一个工作表。
  - 参数 `file_path`：Excel（.xlsx）文件的路径。
  - 参数 `sheet_name`：要打开的工作表的名称。如果没有提供，则返回活动工作表。
  - 返回值：从Excel工作簿中请求的工作表。

- `find_header`: 确定给定Excel工作表是否有标题行。
  - 参数 `file_path`、`max_rows_to_check`、`sheet_name` 类似上面的描述。
  - 返回值：如果找到标题行，第一项为True，否则为False；第二项为标题行的索引，否则为-1。

- `has_header`: 返回 `find_header` 函数是否找到标题行的True/False。
  - 参数类似上面的描述。
  - 返回值：如果找到标题行，则为True，否则为False。

- `get_headers`: 读取Excel（.xlsx）文件的标题。
  - 参数 `file_path`、`sheet_name` 类似上面的描述。
  - 返回值：CSV文件中的列名称列表，以及标题行的索引。

- `read_row`: 从Excel（.xlsx）文件中获取一行的内容。
  - 参数 `file_path`、`row_idx`、`sheet_name` 类似上面的描述。
  - 返回值：一行的内容。
  - 如果文件为空或没有足够的行，则会引发 `ValueError` 异常。

- `read_to_dict`: 解析Excel文件为字典列表。
  - 参数 `file_path`、`start_row`、`sheet_name` 类似上面的描述。
  - 返回值：代表Excel文件的每一行的字典列表。

- `read_to_dict_n_rows`: 解析Excel文件的n行为字典列表。
  - 参数 `file_path`、`n`、`start_row`、`sheet_name` 类似上面的描述。
  - 返回值：代表Excel文件的每一行的字典列表。

这些函数在GeneWeaver程序中的作用是帮助处理和分析基因组数据。例如，`get_sheet_names` 和 `get_sheet` 可以帮助程序加载Excel工作表并读取其内容，`find_header` 和 `has_header` 用于识别数据的结构，以便正确解析，`get_headers`、`read_row`、`read_to_dict` 和 `read_to_dict_n_rows` 则用于将Excel文件中的数据转换为程序可以处理的格式。
"""
from typing import Dict, List, Optional, Tuple, Union
from zipfile import BadZipFile

from geneweaver.core.types import StringOrPath
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet


def get_sheet_names(file_path: StringOrPath) -> List[str]:
    """Retrieve the names of all sheets in an Excel file.

    :param file_path: The path to the Excel file.

    :returns: A list of strings representing the names of all sheets in the file.
    """
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
    except (InvalidFileException, BadZipFile) as e:
        raise ValueError(str(e)) from e
    return workbook.sheetnames


def get_sheet(file_path: StringOrPath, sheet_name: str = None) -> Worksheet:
    """Load a workbook and return a sheet.

    This function opens an Excel workbook and returns the specified sheet if a sheet
    name is provided.

    If no sheet name is provided, it returns the active sheet.

    :param file_path: Path to the Excel (.xlsx) file.
    :param sheet_name: The name of the sheet to open. If not provided, the active sheet
    is returned.

    :returns: The requested worksheet from the Excel workbook.
    """
    workbook = load_workbook(filename=file_path, read_only=True)
    if sheet_name:
        return workbook[sheet_name]
    else:
        return workbook.active


def find_header(
    file_path: StringOrPath, max_rows_to_check: int = 5, sheet_name: str = None
) -> Tuple[bool, int]:
    """Determine if a given Excel worksheet has a header row.

    This function will check up to 'max_rows_to_check' rows from the top to find a
    header row. It assumes that a row can be interpreted as a header if none of its
    values are numeric, if it has more than one column, and if the following row has the
    same number of columns that it does.

    :param file_path: Path to the Excel (.xlsx) file.
    :param max_rows_to_check: The number of rows to check from the top to find a header
    row. Default is 5.
    :param sheet_name: The name of the sheet to open. If not provided, the active sheet
    is returned.

    :returns:
        First index - True if a header row is found, False otherwise.
        Second index - The index of the header row if found, otherwise -1.
    """
    sheet = get_sheet(file_path, sheet_name)
    for i, row in enumerate(sheet.iter_rows(max_row=max_rows_to_check), start=1):
        values = [cell.value for cell in row]
        if (
            len(values) > 1
            and all(value is not None for value in values)
            and all(not isinstance(value, (int, float)) for value in values)
        ):
            next_values = [cell.value for cell in sheet[i + 1]]
            if next_values and len(values) == len(next_values):
                return True, i - 1
    return False, -1


def has_header(
    file_path: str, max_rows_to_check: int = 5, sheet_name: str = None
) -> bool:
    """Summary function to return true/false for if a header is found by `find_header`.

    :param file_path: Path to the Excel (.xlsx) file.
    :param max_rows_to_check: The number of rows to check from the top to find a
    header row.

    :returns:  True if a header row is found, False otherwise.
    """
    return find_header(file_path, max_rows_to_check, sheet_name)[0]


def get_headers(
    file_path: StringOrPath, sheet_name: Optional[str] = None
) -> Tuple[List[str], int]:
    """Read the headers of an Excel (.xlsx) file.

    Read the first row from an Excel (.xlsx) file and return it as a list of strings.

    Assumes that the first row of the Excel file is the header row.

    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to read from. If not provided, the function
    will read from the active sheet.

    :returns: List of column names from the CSV file, and the index of the
    header row.
    """
    headers = []
    file_has_header, header_idx = find_header(file_path, sheet_name=sheet_name)
    if file_has_header:
        headers = read_row(file_path, header_idx, sheet_name=sheet_name)
    return headers, header_idx


def read_row(file_path: str, row_idx: int = 0, sheet_name: str = None) -> List[str]:
    """Get the contents of a row from an Exel (.xlsx) file.

    :param file_path: The file path to the CSV file.
    :param row_idx: The index of the row from which to read the headers. Defaults to 0.
    :param sheet_name: Name of the sheet to read from. If not provided, the function
    will read from the active sheet.

    :returns: The contents of a row

    :raises ValueError: If the file is empty or does not contain enough rows.
    """
    sheet = get_sheet(file_path, sheet_name)
    row_idx += 1
    try:
        return [cell.value for cell in sheet[row_idx]]
    except IndexError as e:
        raise ValueError(f"File does not contain a row at index {row_idx}") from e


def read_to_dict(
    file_path: StringOrPath, start_row: int = 0, sheet_name: Optional[str] = None
) -> List[Dict[str, Union[str, int]]]:
    """Parse an Excel file into a list of dictionaries.

    Parse an Excel file into a list of dictionaries, with keys as column names and
    values as column values.

    :param file_path: The file path to the Excel file.
    :param start_row: The row number to start reading from (0-indexed).
                               This row will be used as the header.
                               Defaults to 0.
    :param sheet_name: Name of the sheet to read from. If not provided, the function
    will read from the active sheet.

    :returns: A list of dictionaries, where each dictionary
    represents a row from the Excel file.
    """
    headers = read_row(file_path, start_row, sheet_name=sheet_name)
    sheet = get_sheet(file_path, sheet_name)
    data = []
    for row in sheet.iter_rows(min_row=start_row + 2):
        row_data = {h: cell.value for h, cell in zip(headers, row)}  # noqa: B905
        data.append(row_data)
    return data


def read_to_dict_n_rows(
    file_path: str, n: int, start_row: int = 0, sheet_name: str = None
) -> List[Dict[str, str]]:
    """Parse n lines of an Excel file into a list of dictionaries.

    Parse an Excel file into a list of dictionaries, with keys as column names and
    values as column values. This function will only return the first n rows of data.

    :param file_path: The file path to the Excel file.
    :param n: The number of rows of data to return.
    :param start_row: The row number to start reading from (0-indexed).
                           This row will be used as the header.
                           Defaults to 0.
    :param sheet_name: Name of the sheet to read from. If not provided, the function
    will read from the active sheet.

    :returns: A list of dictionaries, where each dictionary represents
    a row from the Excel file.
    """
    headers = read_row(file_path, start_row, sheet_name=sheet_name)
    sheet = get_sheet(file_path, sheet_name)
    data = []
    for row in sheet.iter_rows(min_row=start_row + 2, max_row=start_row + 1 + n):
        row_data = {h: cell.value for h, cell in zip(headers, row)}  # noqa: B905
        data.append(row_data)
    return data
